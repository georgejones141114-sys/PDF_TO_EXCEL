import io
import os
import re

import openpyxl

from app import app


FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "sample_rse_report.pdf")


def test_convert_route_offers_optional_workbook_actions():
    app.config.update(TESTING=True)
    with app.test_client() as client, open(FIXTURE, "rb") as pdf:
        response = client.post(
            "/convert",
            data={"pdf_file": (pdf, "client-report.pdf")},
            content_type="multipart/form-data",
        )

    assert response.status_code == 200
    assert "Download this workbook" in response.text
    assert "Add to master sheet" in response.text
    assert "Workbook preview" in response.text
    assert "STOCK" in response.text
    assert "BOK" in response.text
    assert "Content-Disposition" not in response.headers

    token = re.search(r"/download/([a-f0-9]+)", response.text).group(1)
    with app.test_client() as client:
        add_response = client.post(f"/add-to-master/{token}")
        assert add_response.status_code == 200
        assert "was added to your master workbook" in add_response.text

        master_response = client.get("/download-master")

    assert master_response.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(master_response.data))
    assert workbook.sheetnames[0] == "MASTER"
    assert workbook["MASTER"]["B2"].value == "client-report.pdf"
    assert workbook["MASTER"]["C2"].value is not None


def test_download_converted_workbook_remains_available():
    app.config.update(TESTING=True)
    with app.test_client() as client, open(FIXTURE, "rb") as pdf:
        response = client.post(
            "/convert",
            data={"pdf_file": (pdf, "client-report.pdf")},
            content_type="multipart/form-data",
        )
        token = re.search(r"/download/([a-f0-9]+)", response.text).group(1)
        download = client.get(f"/download/{token}")

    assert download.status_code == 200
    assert download.headers["Content-Disposition"].startswith("attachment;")
    workbook = openpyxl.load_workbook(io.BytesIO(download.data))
    assert workbook.sheetnames[0] == "STOCK"