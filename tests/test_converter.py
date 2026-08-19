import os
import io
import openpyxl
import pytest

from converter import convert, ReportParseError

FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "sample_rse_report.pdf")


@pytest.fixture
def sample_pdf_bytes():
    with open(FIXTURE, "rb") as f:
        return f.read()


def test_convert_produces_expected_sheets(sample_pdf_bytes):
    xlsx_bytes = convert(sample_pdf_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == [
        "STOCK",
        "MARKET STATS",
        "EXCHANGE RATE",
        "EQUITIES MARKET",
        "BONDS_GOV+CORP",
        "BONDS TRADES",
        "WEEKLY REPORT",
    ]


def test_equities_market_rows(sample_pdf_bytes):
    xlsx_bytes = convert(sample_pdf_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["EQUITIES MARKET"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("STOCK", "CLOSING", "VOLUME", "VALUE")
    bok_row = next(r for r in rows[1:] if r[0] == "BOK")
    assert bok_row == ("BOK", 660, 10700, 7062000)


def test_market_stats_values(sample_pdf_bytes):
    xlsx_bytes = convert(sample_pdf_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    stats = dict((r[0], r[1]) for r in wb["MARKET STATS"].iter_rows(values_only=True, min_row=2))
    assert stats["RSI"] == 213.74
    assert stats["ALSI"] == 259.11
    assert stats["Bond market today (FRW)"] == 1717401000


def test_bond_coupon_rates_are_numeric(sample_pdf_bytes):
    xlsx_bytes = convert(sample_pdf_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["BONDS_GOV+CORP"]

    coupon_values = [row[1] for row in ws.iter_rows(values_only=True, min_row=2)]

    assert coupon_values
    assert all(isinstance(value, (int, float)) for value in coupon_values)


def test_weekly_report_total(sample_pdf_bytes):
    xlsx_bytes = convert(sample_pdf_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    rows = list(wb["WEEKLY REPORT"].iter_rows(values_only=True))
    total_row = next(r for r in rows if r[0] == "TOTAL")
    assert total_row == ("TOTAL", 682300, 301915700, 23)


def test_rejects_non_rse_pdf():
    reportlab = pytest.importorskip("reportlab")  # dev-only dependency, see requirements-dev.txt
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf)
    c.drawString(72, 720, "Just a random document, not a market report.")
    c.save()

    with pytest.raises(ReportParseError):
        convert(buf.getvalue())
