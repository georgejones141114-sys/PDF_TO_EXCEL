"""
converter.py
------------
Parses a Rwanda Stock Exchange (RSE) weekly/daily "Market Report" PDF and
produces a styled Excel (.xlsx) workbook with one sheet per data table:

    STOCK            - security / closing price / volume / value
    MARKET STATS     - RSI, ALSI, equity turnover, bond turnover, market cap
    EXCHANGE RATE    - buying / selling exchange rates
    EQUITIES MARKET  - full equities trading table
    BONDS_GOV+CORP   - government + corporate bonds trading table
    BONDS TRADES     - bonds that actually traded today (best-effort, see note)
    WEEKLY REPORT    - this week's daily volume / value / deals


NOTE on "BONDS TRADES": the source PDF only reports a *total* traded
volume per bond for the day, not a deal-by-deal breakdown. The original
sample workbook had deal-by-deal rows that were reconstructed by hand.
This converter instead emits one row per bond that traded (best-effort),
clearly documented so nobody mistakes it for the manually reconstructed
version.
"""

import re
import io
from copy import copy
from datetime import datetime
from dataclasses import dataclass, field

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# STYLING
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
DATA_FONT = Font(name="Calibri", bold=False, size=10)
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")


class ReportParseError(Exception):
    """Raised when the uploaded PDF doesn't look like an RSE market report."""


def _num(s):
    """Best-effort string -> int/float, keeping % / signed strings as text."""
    if not isinstance(s, str):
        return s
    s2 = s.replace(",", "").strip()
    if s2 in ("-", ""):
        return 0
    try:
        if "%" in s2 or s2.startswith("+"):
            return s2
        if "." in s2:
            return float(s2)
        return int(s2)
    except ValueError:
        return s


def _write_sheet(wb, name, headers, rows, numeric_cols):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HEADER_FONT
        c.border = BORDER
        c.fill = HEADER_FILL

    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            v = _num(val) if j in numeric_cols else val
            c = ws.cell(row=i, column=j, value=v)
            c.font = DATA_FONT
            c.border = BORDER

    for j, h in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        max_len = max([len(str(h))] + [len(str(r[j - 1])) for r in rows]) if rows else len(str(h))
        ws.column_dimensions[col_letter].width = max_len + 3

    ws.freeze_panes = "A2"
    return ws


def _combine_security_status(row):
    security, status = row[2], row[1]
    return f"{security} ({status})" if status else security


# EXTRACTION
ISIN_RE = re.compile(r"^(RW[0-9A-Z]{8,}|KE[0-9A-Z]{8,}|ZAE[0-9A-Z]{8,}|MGMRW|PRERW|AMSRW)\s")
EQUITY_ROW_RE = re.compile(r"^(RW|KE|ZAE)\w+ [A-Z]+ ")
WEEKLY_ROW_RE = re.compile(r"^(\d\d-|TOTAL)")
FX_RE = re.compile(r"^(USD|KES|UGS|UGX|BIF|TZS|ZAR|SAR)\s+([\d,]+\.\d+)\s+([\d,]+\.\d+)\s+([\d,]+\.\d+)$")


def _parse_bond_line(l):
    toks = l.split()
    if toks[1] == "Re-opened":
        status, isin, rest = "Re-opened", toks[0], toks[2:]
    else:
        status, isin, rest = "", toks[0], toks[1:]
    if len(rest) != 8:
        return None
    security, maturity, coupon, close_p, prev_p, bids, offers, traded = rest
    return [isin, status, security, maturity, coupon, close_p, prev_p, bids, offers, traded]


def _page_text_column_safe(page):
    """
    Extract a page's text without letting pdfplumber interleave lines from
    two side-by-side prose columns (the report's cover page reads "Market
    overview" on the left and "Closing bell" on the right - if extracted
    naively, sentences from each column get shuffled together line by
    line). We crop left/right halves and concatenate them so regexes never
    see a sentence from one column broken up by a line from the other.

    This is only applied to pages that are genuinely prose-in-two-columns
    (short lines, lots of them on both halves) - wide data tables (like the
    equities/bonds tables) must NOT be column-split, or rows get cut in
    half. We detect prose columns by requiring many lines on both halves
    AND a low average line length (data table rows are long).
    """
    plain = page.extract_text() or ""

    # Only the "Market overview" / "Closing bell" cover section uses a real
    # two-column prose layout in RSE reports. Data-table pages (equities,
    # bonds, weekly report) are full-width and must never be column-split,
    # or table rows get sliced in half. Detect the overview page by its
    # section markers rather than by page number, so this still works if a
    # future report shuffles page order.
    if "market overview" not in plain.lower() or "closing bell" not in plain.lower():
        return plain

    w = page.width
    left = page.crop((0, 0, w / 2, page.height)).extract_text() or ""
    right = page.crop((w / 2, 0, w, page.height)).extract_text() or ""
    if left.strip() and right.strip():
        return left + "\n" + right
    return plain


def _extract_lines(pdf):
    """All lines from every page, in order, plus the raw per-page text."""
    pages_text = [_page_text_column_safe(p) for p in pdf.pages]
    all_lines = []
    for t in pages_text:
        all_lines.extend(t.split("\n"))
    return pages_text, all_lines


def _extract_equities(all_lines):
    rows = []
    for l in all_lines:
        if EQUITY_ROW_RE.match(l):
            toks = l.split()
            if len(toks) == 11:
                rows.append(toks)
    return rows


def _extract_bonds(all_lines):
    gov_rows, corp_rows = [], []
    corp_started = False
    for l in all_lines:
        if l.strip().lower().startswith("b. corporate bonds"):
            corp_started = True
            continue
        if ISIN_RE.match(l):
            parsed = _parse_bond_line(l)
            if parsed is None:
                continue
            (corp_rows if corp_started else gov_rows).append(parsed)
    return gov_rows, corp_rows


def _extract_weekly(all_lines):
    rows = []
    for raw in all_lines:
        l = raw.replace("04- August-26", "04-August-26")
        # normalise "DD- Month-YY" -> "DD-Month-YY" generally (stray space after day)
        l = re.sub(r"^(\d\d)- ", r"\1-", l)
        if WEEKLY_ROW_RE.match(l):
            toks = l.split()
            if len(toks) == 10:
                rows.append(toks)
    return rows


def _extract_market_stats(all_lines, full_text):
    stats = {}

    m = re.search(r"RSI\s+([\d.]+)\s+([\d.]+)\s+([+-][\d.]+)\s+([+-][\d.]+)", full_text)
    if m:
        stats["RSI"] = float(m.group(2))
    m = re.search(r"ALSI\s+([\d.]+)\s+([\d.]+)\s+([+-][\d.]+)\s+([+-][\d.]+)", full_text)
    if m:
        stats["ALSI"] = float(m.group(2))

    m = re.search(r"Equity Turnover\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)", full_text)
    if m:
        stats["Equity turnover"] = _num(m.group(2))

    m = re.search(r"Frw\s+([\d,]+)\s+worth of\s*\n?\s*bonds traded in (\d+) deals", full_text)
    if m:
        stats["Bond market today (FRW)"] = _num(m.group(1))

    m = re.search(r"Market Capitalization \(Frw\)\s*\n?\s*([\d,]+)", full_text)
    if m:
        stats["Market capitalization (FRW)"] = _num(m.group(1))

    order = ["RSI", "ALSI", "Equity turnover", "Bond market today (FRW)", "Market capitalization (FRW)"]
    return [[k, stats[k]] for k in order if k in stats]


def _extract_fx(all_lines):
    rows = []
    code_map = {"UGS": "UGX", "ZAR": "SAR"}  # normalise report code -> ISO-ish label if desired
    for l in all_lines:
        m = FX_RE.match(l.strip())
        if m:
            code, sell, buy, avg = m.groups()
            rows.append([code, _num(buy), _num(sell)])
    return rows


def _extract_bond_trades(gov_rows, corp_rows):
    """Best-effort: one row per bond that traded today (see module docstring)."""
    rows = []
    for category, bucket in (("TREASURY", gov_rows), ("CORPORATE", corp_rows)):
        for r in bucket:
            isin, status, security, maturity, coupon, close_p, prev_p, bids, offers, traded = r
            traded_val = _num(traded)
            if traded_val in (0, "0", 0.0):
                continue
            close_v, prev_v = _num(close_p), _num(prev_p)
            try:
                change = round(float(close_v) - float(prev_v), 4)
            except (TypeError, ValueError):
                change = ""
            name = _combine_security_status([isin, status, security])
            rows.append([name, category, traded_val, prev_v, close_v, change])
    return rows


# --------------------------------------------------------------------------
# Public entry point
# --------------------------------------------------------------------------
def convert(pdf_bytes: bytes, source_filename="", added_at=None) -> bytes:
    """Convert an RSE market report PDF into an .xlsx workbook with a master sheet."""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        pages_text, all_lines = _extract_lines(pdf)

    full_text = "\n".join(pages_text)

    if "RWANDA STOCK EXCHANGE" not in full_text.upper():
        raise ReportParseError(
            "This doesn't look like a Rwanda Stock Exchange market report PDF."
        )

    equities_rows = _extract_equities(all_lines)
    gov_rows, corp_rows = _extract_bonds(all_lines)
    weekly_rows = _extract_weekly(all_lines)

    if not equities_rows and not gov_rows and not weekly_rows:
        raise ReportParseError(
            "Couldn't find any recognisable equities, bonds, or weekly tables in this PDF. "
            "The layout may differ from the expected RSE report format."
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    # STOCK (security / closing / volume / value) - derived from the equities table
    eq_keep = [1, 6, 9, 10]
    eq_headers = ["STOCK", "CLOSING", "VOLUME", "VALUE"]
    eq_filtered = [[row[i] for i in eq_keep] for row in equities_rows]
    _write_sheet(wb, "STOCK", eq_headers, eq_filtered, numeric_cols={2, 3, 4})

    # MARKET STATS
    stats_rows = _extract_market_stats(all_lines, full_text)
    _write_sheet(wb, "MARKET STATS", ["INDICATORS", "CLOSING"], stats_rows, numeric_cols={2})

    # EXCHANGE RATE
    fx_rows = _extract_fx(all_lines)
    _write_sheet(wb, "EXCHANGE RATE", ["CURRENCY CODE", "BUYING VALUE", "SELLING VALUE"], fx_rows, numeric_cols={2, 3})

    # EQUITIES MARKET (full table)
    _write_sheet(wb, "EQUITIES MARKET", eq_headers, eq_filtered, numeric_cols={2, 3, 4})

    # BONDS_GOV+CORP
    bond_headers = ["BOND", "COUPON RATE", "CLOSE. PRICE", "PREV. PRICE", "BOND TRADED"]
    gov_filtered = [[_combine_security_status(row), row[4], row[5], row[6], row[9]] for row in gov_rows]
    corp_filtered = [[_combine_security_status(row), row[4], row[5], row[6], row[9]] for row in corp_rows]
    _write_sheet(wb, "BONDS_GOV+CORP", bond_headers, gov_filtered + corp_filtered, numeric_cols={3, 4, 5})

    # BONDS TRADES (best-effort, one row per bond - see docstring)
    trades_rows = _extract_bond_trades(gov_rows, corp_rows)
    _write_sheet(
        wb, "BONDS TRADES",
        ["BOND", "CATEGORY", "VOLUME", "PREVIOUS", "CLOSING", "CHANGE"],
        trades_rows, numeric_cols={3, 4, 5, 6},
    )

    # WEEKLY REPORT (this week's volume / value / deals only, per date)
    weekly_keep = [4, 5, 6, 7]
    weekly_headers = ["This Week Date", "Volume", "Value (Frw)", "Deals"]
    weekly_filtered = [[row[i] for i in weekly_keep] for row in weekly_rows]
    _write_sheet(wb, "WEEKLY REPORT", weekly_headers, weekly_filtered, numeric_cols={2, 3, 4})

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def add_to_master(master_bytes, report_bytes, source_filename, added_at=None):
    """Add a converted report's sheets to a master workbook."""
    added_at = added_at or datetime.now()

    report_wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
    if master_bytes:
        master_wb = openpyxl.load_workbook(io.BytesIO(master_bytes))
    else:
        master_wb = openpyxl.Workbook()
        master_wb.active.title = "MASTER"
        _write_sheet(master_wb, "MASTER LOG", [], [], numeric_cols=set())
        del master_wb["MASTER LOG"]
        master = master_wb["MASTER"]
        for column, header in enumerate(["SPREADSHEET", "SOURCE PDF", "ADDED AT"], start=1):
            cell = master.cell(row=1, column=column, value=header)
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.fill = HEADER_FILL
        master.freeze_panes = "A2"

    master = master_wb["MASTER"]
    row = master.max_row + 1
    source_stem = re.sub(r"[^A-Za-z0-9]+", "_", source_filename.rsplit(".", 1)[0]).strip("_") or "Report"

    for sheet in report_wb.worksheets:
        if sheet.title == "MASTER":
            continue
        base_name = f"{source_stem} - {sheet.title}"[:31]
        sheet_name = base_name
        suffix = 2
        while sheet_name in master_wb.sheetnames:
            suffix_text = f" {suffix}"
            sheet_name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1

        copied = master_wb.create_sheet(sheet_name)
        copied.sheet_view.showGridLines = False
        for source_row in sheet.iter_rows():
            for source_cell in source_row:
                target_cell = copied.cell(
                    row=source_cell.row,
                    column=source_cell.column,
                    value=source_cell.value,
                )
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
        for column, dimension in sheet.column_dimensions.items():
            copied.column_dimensions[column].width = dimension.width
        copied.freeze_panes = sheet.freeze_panes

        master.cell(row=row, column=1, value=sheet_name)
        master.cell(row=row, column=2, value=source_filename)
        master.cell(row=row, column=3, value=added_at)
        for cell in master[row]:
            cell.border = BORDER
        master.cell(row=row, column=3).number_format = "yyyy-mm-dd hh:mm:ss"
        row += 1

    for column in range(1, 4):
        master.column_dimensions[get_column_letter(column)].width = [35, 30, 22][column - 1]

    out = io.BytesIO()
    master_wb.save(out)
    return out.getvalue()
